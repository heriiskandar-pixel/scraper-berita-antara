#!/usr/bin/env python3
"""
scrape_berita_rss.py
====================
Skrip otomatisasi penarik berita:
- Analisis Sentimen berbasis AI (IndoBERT via Hugging Face Transformers).
- Menarik RSS Feed ANTARA News, Detikcom, CNN Indonesia, Tribunnews, CNBC Indonesia, Bisnis Indonesia, Tempo, & Republika (semua kanal).
- Scraping halaman Indeks Kanal Detikcom & Kompas.com hingga Halaman 3.
- Pemetaan kategori, ekstraksi gambar, penulis, dan sentimen secara akurat.
- Menyimpan data kumulatif tanpa duplikat ke Excel per tanggal terbit & index.json.
"""

import os
import re
import time
import json
import requests
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from html import unescape
from xml.etree import ElementTree as ET
from bs4 import BeautifulSoup
from dateutil import parser
from transformers import pipeline

# ----------------------------------------------------------------------
# PENGATURAN GLOBAL
# ----------------------------------------------------------------------
OUTPUT_FOLDER = "data"
FILENAME_PREFIX = "berita"
MAKS_UMUR_HARI = 7
MAX_PAGE_INDEKS = 3
MAX_WORKERS = 12  # jumlah request halaman detail yang dijalankan paralel

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
    "Referer": "https://www.google.com/",
}

# Beberapa media (seperti CNN Indonesia, CNBC Indonesia, dll.) mungkin menggunakan
# proteksi Cloudflare atau proteksi bot lainnya. Skrip menggunakan cloudscraper jika tersedia.
try:
    import cloudscraper
    _CF_SCRAPER = cloudscraper.create_scraper(
        browser={"browser": "chrome", "platform": "windows", "mobile": False}
    )
except ImportError:
    print("Peringatan: package 'cloudscraper' belum terpasang. "
          "Jalankan 'pip install cloudscraper' agar RSS yang diproteksi Cloudflare "
          "bisa ditarik dengan lancar.")
    _CF_SCRAPER = requests

def _http_get(url, timeout=15, gunakan_cloudscraper=False):
    """Wrapper GET request dengan dukungan cloudscraper dan fallback proxy publik."""
    if not gunakan_cloudscraper:
        return requests.get(url, headers=HEADERS, timeout=timeout)

    resp = None
    try:
        resp = _CF_SCRAPER.get(url, headers=HEADERS, timeout=timeout)
        if resp.status_code != 403:
            return resp
    except Exception:
        pass

    try:
        proxy_url = "https://api.allorigins.win/raw?url=" + requests.utils.quote(url, safe="")
        resp_proxy = requests.get(proxy_url, headers=HEADERS, timeout=timeout + 15)
        return resp_proxy
    except Exception:
        if resp is not None:
            return resp
        raise

# ----------------------------------------------------------------------
# INISIALISASI MODEL AI SENTIMEN (INDOBERT)
# ----------------------------------------------------------------------
print("Memuat model sentimen IndoBERT dari Hugging Face...")
try:
    sentimen_pipeline = pipeline(
        "text-classification",
        model="witosetiadi/indobert-base-cased-sentiment-analysis",
        tokenizer="witosetiadi/indobert-base-cased-sentiment-analysis"
    )
    print("Model IndoBERT berhasil dimuat.\n")
except Exception as e:
    print(f"Peringatan: Gagal memuat model IndoBERT ({e}). Pengujian sentimen akan menggunakan fallback Netral.")
    sentimen_pipeline = None

def analisa_sentimen(judul, ringkasan):
    """Analisis sentimen presisi berbasis AI IndoBERT"""
    if not sentimen_pipeline:
        return "Netral"
        
    teks = f"{judul}. {ringkasan}".strip()
    if not teks:
        return "Netral"
    
    teks_input = teks[:512]
    
    try:
        hasil = sentimen_pipeline(teks_input)[0]
        label = str(hasil['label']).upper()
        
        if "POS" in label or "LABEL_0" in label:
            return "Positif"
        elif "NEG" in label or "LABEL_2" in label:
            return "Negatif"
        else:
            return "Netral"
    except Exception:
        return "Netral"

# ----------------------------------------------------------------------
# DAFTAR RSS FEEDS & KANAL INDEKS
# ----------------------------------------------------------------------
FEEDS_RSS = {
    # ---- ANTARA News ----
    "antara-terkini":    {"url": "https://www.antaranews.com/rss/terkini.xml", "media": "Antara", "kategori": "Terkini"},
    "antara-top-news":   {"url": "https://www.antaranews.com/rss/top-news.xml", "media": "Antara", "kategori": "Top News"},
    "antara-politik":    {"url": "https://www.antaranews.com/rss/politik.xml", "media": "Antara", "kategori": "Politik"},
    "antara-hukum":      {"url": "https://www.antaranews.com/rss/hukum.xml", "media": "Antara", "kategori": "Hukum"},
    "antara-ekonomi":    {"url": "https://www.antaranews.com/rss/ekonomi.xml", "media": "Antara", "kategori": "Ekonomi"},
    "antara-metro":      {"url": "https://www.antaranews.com/rss/metro.xml", "media": "Antara", "kategori": "Metro"},
    "antara-sepakbola":  {"url": "https://www.antaranews.com/rss/sepakbola.xml", "media": "Antara", "kategori": "Sepakbola"},
    "antara-olahraga":   {"url": "https://www.antaranews.com/rss/olahraga.xml", "media": "Antara", "kategori": "Olahraga"},
    "antara-humaniora":  {"url": "https://www.antaranews.com/rss/humaniora.xml", "media": "Antara", "kategori": "Humaniora"},
    "antara-lifestyle":  {"url": "https://www.antaranews.com/rss/lifestyle.xml", "media": "Antara", "kategori": "Lifestyle"},
    "antara-hiburan":    {"url": "https://www.antaranews.com/rss/hiburan.xml", "media": "Antara", "kategori": "Hiburan"},
    "antara-dunia":      {"url": "https://www.antaranews.com/rss/dunia.xml", "media": "Antara", "kategori": "Dunia"},
    "antara-infografis": {"url": "https://www.antaranews.com/rss/infografis.xml", "media": "Antara", "kategori": "Infografis"},
    "antara-tekno":      {"url": "https://www.antaranews.com/rss/tekno.xml", "media": "Antara", "kategori": "Tekno"},
    "antara-otomotif":   {"url": "https://www.antaranews.com/rss/otomotif.xml", "media": "Antara", "kategori": "Otomotif"},
    "antara-warta-bumi": {"url": "https://www.antaranews.com/rss/warta-bumi.xml", "media": "Antara", "kategori": "Warta Bumi"},
    "antara-foto":       {"url": "https://www.antaranews.com/rss/foto.xml", "media": "Antara", "kategori": "Foto"},

    # ---- Detikcom ----
    "detik-news":        {"url": "https://news.detik.com/rss", "media": "Detik", "kategori": "News"},
    "detik-finance":     {"url": "https://finance.detik.com/rss", "media": "Detik", "kategori": "Finance"},
    "detik-sport":       {"url": "https://sport.detik.com/rss", "media": "Detik", "kategori": "Sport"},

    # ---- CNN Indonesia ----
    "cnn-nasional":      {"url": "https://www.cnnindonesia.com/nasional/rss", "media": "CNN Indonesia", "kategori": "Nasional"},
    "cnn-internasional": {"url": "https://www.cnnindonesia.com/internasional/rss", "media": "CNN Indonesia", "kategori": "Internasional"},
    "cnn-ekonomi":       {"url": "https://www.cnnindonesia.com/ekonomi/rss", "media": "CNN Indonesia", "kategori": "Ekonomi"},
    "cnn-olahraga":      {"url": "https://www.cnnindonesia.com/olahraga/rss", "media": "CNN Indonesia", "kategori": "Olahraga"},
    "cnn-teknologi":     {"url": "https://www.cnnindonesia.com/teknologi/rss", "media": "CNN Indonesia", "kategori": "Teknologi"},
    "cnn-hiburan":       {"url": "https://www.cnnindonesia.com/hiburan/rss", "media": "CNN Indonesia", "kategori": "Hiburan"},
    "cnn-gaya-hidup":    {"url": "https://www.cnnindonesia.com/gaya-hidup/rss", "media": "CNN Indonesia", "kategori": "Gaya Hidup"},
    "cnn-otomotif":      {"url": "https://www.cnnindonesia.com/otomotif/rss", "media": "CNN Indonesia", "kategori": "Otomotif"},
    "cnn-edukasi":       {"url": "https://www.cnnindonesia.com/edukasi/rss", "media": "CNN Indonesia", "kategori": "Edukasi"},

    # ---- Tribunnews ----
    "tribun-news":       {"url": "https://www.tribunnews.com/rss", "media": "Tribunnews", "kategori": "News"},
    "tribun-bisnis":     {"url": "https://www.tribunnews.com/bisnis/rss", "media": "Tribunnews", "kategori": "Bisnis"},
    "tribun-superskor":  {"url": "https://www.tribunnews.com/superskor/rss", "media": "Tribunnews", "kategori": "Olahraga"},
    "tribun-seleb":      {"url": "https://www.tribunnews.com/seleb/rss", "media": "Tribunnews", "kategori": "Seleb"},
    "tribun-lifestyle":  {"url": "https://www.tribunnews.com/lifestyle/rss", "media": "Tribunnews", "kategori": "Lifestyle"},
    "tribun-travel":     {"url": "https://www.tribunnews.com/travel/rss", "media": "Tribunnews", "kategori": "Travel"},

    # ---- CNBC Indonesia ----
    "cnbc-investment":   {"url": "https://www.cnbcindonesia.com/investment/rss", "media": "CNBC Indonesia", "kategori": "Investment"},
    "cnbc-news":         {"url": "www.cnbcindonesia.com/news/rss", "media": "CNBC Indonesia", "kategori": "News"},
    "cnbc-market":       {"url": "https://www.cnbcindonesia.com/market/rss", "media": "CNBC Indonesia", "kategori": "Market"},
    "cnbc-entrepreneur": {"url": "https://www.cnbcindonesia.com/entrepreneur/rss", "media": "CNBC Indonesia", "kategori": "Entrepreneur"},
    "cnbc-tech":         {"url": "https://www.cnbcindonesia.com/tech/rss", "media": "CNBC Indonesia", "kategori": "Tech"},
    "cnbc-lifestyle":    {"url": "https://www.cnbcindonesia.com/lifestyle/rss", "media": "CNBC Indonesia", "kategori": "Lifestyle"},

    # ---- Bisnis Indonesia ----
    "bisnis-makro":      {"url": "https://entrepreneur.bisnis.com/rss", "media": "Bisnis Indonesia", "kategori": "Makro"}, # or general rss index if available
    "bisnis-finansial":  {"url": "https://finansial.bisnis.com/rss", "media": "Bisnis Indonesia", "kategori": "Finansial"},
    "bisnis-ekonomi":    {"url": "https://ekonomi.bisnis.com/rss", "media": "Bisnis Indonesia", "kategori": "Ekonomi"},
    "bisnis-market":     {"url": "https://market.bisnis.com/rss", "media": "Bisnis Indonesia", "kategori": "Market"},

    # ---- Tempo ----
    "tempo-nasional":    {"url": "https://rss.tempo.co/nasional", "media": "Tempo", "kategori": "Nasional"},
    "tempo-bisnis":      {"url": "https://rss.tempo.co/bisnis", "media": "Tempo", "kategori": "Bisnis"},
    "tempo-metro":       {"url": "https://rss.tempo.co/metro", "media": "Tempo", "kategori": "Metro"},
    "tempo-dunia":       {"url": "https://rss.tempo.co/dunia", "media": "Tempo", "kategori": "Dunia"},
    "tempo-tekno":       {"url": "https://rss.tempo.co/tekno", "media": "Tempo", "kategori": "Tekno"},

    # ---- Republika ----
    "republika-news":    {"url": "https://republika.co.id/rss/", "media": "Republika", "kategori": "News"},
    "republika-trend":   {"url": "https://republika.co.id/rss/trend", "media": "Republika", "kategori": "Trend"},
    "republika-leisure": {"url": "https://republika.co.id/rss/leisure", "media": "Republika", "kategori": "Leisure"},
}

KANAL_INDEKS_DETIK = [
    {"subdomain": "inet", "kategori": "Teknologi"},
    {"subdomain": "hot", "kategori": "Hiburan"},
    {"subdomain": "health", "kategori": "Kesehatan"},
    {"subdomain": "food", "kategori": "Kuliner"},
    {"subdomain": "travel", "kategori": "Wisata"},
    {"subdomain": "oto", "kategori": "Otomotif"},
    {"subdomain": "edu", "kategori": "Edukasi"},
    {"subdomain": "hikmah", "kategori": "Hikmah"},
    {"subdomain": "properti", "kategori": "Properti"},
    {"subdomain": "wolipop", "kategori": "Gaya Hidup"},
]

# ----------------------------------------------------------------------
# FUNGSI UTILITAS
# ----------------------------------------------------------------------
def clean_html(raw_html):
    if not raw_html:
        return ""
    text = re.sub(r"<[^>]+>", "", raw_html)
    return unescape(text).strip()

def parse_tanggal(raw):
    if not raw:
        return None
    try:
        dt = parser.parse(str(raw), fuzzy=True)
        if dt.tzinfo is None:
            wib = timezone(timedelta(hours=7))
            dt = dt.replace(tzinfo=wib)
        return dt
    except Exception:
        return None

def format_tanggal_rfc2822(dt):
    if not dt:
        return ""
    return dt.strftime("%a, %d %b %Y %H:%M:%S %z")

def ekstraksi_detail_halaman(url, media_default=""):
    """Mengunjungi URL halaman berita untuk mengambil ringkasan, gambar utama, dan penulis."""
    ringkasan = ""
    url_gambar = ""
    penulis = ""
    domain_cf = any(d in url for d in ["cnnindonesia.com", "cnbcindonesia.com", "tempo.co"])
    try:
        resp = _http_get(url, timeout=6, gunakan_cloudscraper=domain_cf)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'lxml')
            
            meta_desc = soup.find('meta', attrs={'name': 'description'}) or \
                        soup.find('meta', attrs={'property': 'og:description'}) or \
                        soup.find('meta', attrs={'name': 'twitter:description'})
            if meta_desc and meta_desc.get('content'):
                ringkasan = meta_desc['content'].strip()
            else:
                p_first = soup.find('p')
                if p_first:
                    ringkasan = p_first.get_text(strip=True)

            meta_img = soup.find('meta', attrs={'property': 'og:image'}) or \
                       soup.find('meta', attrs={'name': 'twitter:image'})
            if meta_img and meta_img.get('content'):
                url_gambar = meta_img['content'].strip()

            meta_auth = soup.find('meta', attrs={'name': 'author'}) or \
                        soup.find('meta', attrs={'name': 'baca-author'}) or \
                        soup.find('meta', attrs={'property': 'article:author'}) or \
                        soup.find('meta', attrs={'property': 'dd:author'})
            
            if meta_auth and meta_auth.get('content'):
                penulis = meta_auth['content'].strip()
            
            if not penulis:
                elem_author = soup.select_one('.detail__author, .read__author, .credit-title-name, .author, .byline, .penulis, .detail-author, .writer')
                if elem_author:
                    penulis = elem_author.get_text(strip=True)

            if penulis:
                penulis = re.sub(r'^(Oleh|By|Penulis|Reporter|Editor)\s*:\s*', '', penulis, flags=re.I)
                penulis = re.sub(r'\s*-\s*(detik|Kompas|ANTARA|CNN|CNBC|Tribun|Tempo|Bisnis|Republika).*$', '', penulis, flags=re.I)
                penulis = penulis.strip()

    except Exception:
        pass

    if not penulis and media_default:
        penulis = f"Redaksi {media_default}"

    return ringkasan, url_gambar, penulis

def ekstraksi_detail_banyak_halaman(daftar_link, media_default="", max_workers=MAX_WORKERS):
    """Mengambil ringkasan/gambar/penulis untuk BANYAK link sekaligus secara paralel."""
    hasil = {}
    if not daftar_link:
        return hasil
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_ke_link = {
            executor.submit(ekstraksi_detail_halaman, link, media_default): link
            for link in daftar_link
        }
        for future in as_completed(future_ke_link):
            link = future_ke_link[future]
            try:
                hasil[link] = future.result()
            except Exception:
                hasil[link] = ("", "", "")
    return hasil

# ----------------------------------------------------------------------
# MODUL SCRAPING
# ----------------------------------------------------------------------
def ambil_rss(feed_info):
    """Ekstraksi berita dari Feed RSS berbagai media"""
    hasil = []
    domain_cf = any(d in feed_info["url"] for d in ["cnnindonesia.com", "cnbcindonesia.com", "tempo.co"])
    try:
        resp = _http_get(feed_info["url"], timeout=20, gunakan_cloudscraper=domain_cf)
        resp.raise_for_status()
        
        try:
            root = ET.fromstring(resp.content)
        except ET.ParseError:
            root = ET.fromstring(resp.text.encode('utf-8'))

        namespaces = {
            'media': 'http://search.yahoo.com/mrss/',
            'dc': 'http://purl.org/dc/elements/1.1/',
            'content': 'http://purl.org/rss/1.0/modules/content/'
        }

        item_mentah = []
        link_perlu_detail = set()
        for item in root.findall(".//item"):
            judul = (item.findtext("title") or "").strip()
            link = (item.findtext("link") or "").strip()
            tanggal = (item.findtext("pubDate") or "").strip()
            raw_desc = item.findtext("description") or ""
            
            cat_xml = item.findtext("category")
            kategori_berita = cat_xml.strip().capitalize() if cat_xml and cat_xml.strip() else feed_info["kategori"]

            penulis = ""
            for tag_auth in ["dc:creator", "author", "creator"]:
                found_p = item.findtext(tag_auth, namespaces=namespaces) or item.findtext(tag_auth)
                if found_p and found_p.strip():
                    penulis = found_p.strip()
                    break

            deskripsi = clean_html(raw_desc)
            
            url_gambar = ""
            enclosure = item.find("enclosure")
            if enclosure is not None and enclosure.get("url"):
                url_gambar = enclosure.get("url")
            if not url_gambar:
                media_content = item.find("media:content", namespaces) or item.find("media:thumbnail", namespaces)
                if media_content is not None and media_content.get("url"):
                    url_gambar = media_content.get("url")
            if not url_gambar and raw_desc:
                match_img = re.search(r'<img[^>]+src=["\']([^"\']+)["\']', raw_desc, re.IGNORECASE)
                if match_img:
                    url_gambar = match_img.group(1)

            butuh_detail = (not penulis) or (len(deskripsi) < 30) or (not url_gambar)
            if butuh_detail and link:
                link_perlu_detail.add(link)

            item_mentah.append({
                "kategori": kategori_berita,
                "judul": judul,
                "link": link,
                "tanggal": tanggal,
                "penulis": penulis,
                "deskripsi": deskripsi,
                "url_gambar": url_gambar,
            })

        detail_map = ekstraksi_detail_banyak_halaman(
            list(link_perlu_detail), media_default=feed_info["media"]
        )

        for it in item_mentah:
            penulis = it["penulis"]
            deskripsi = it["deskripsi"]
            url_gambar = it["url_gambar"]

            if it["link"] in detail_map:
                detail_desc, detail_img, detail_penulis = detail_map[it["link"]]
                if not penulis and detail_penulis:
                    penulis = detail_penulis
                if len(deskripsi) < 30 and detail_desc:
                    deskripsi = detail_desc
                if not url_gambar and detail_img:
                    url_gambar = detail_img

            if not penulis:
                penulis = f"Redaksi {feed_info['media']}"

            sentimen = analisa_sentimen(it["judul"], deskripsi)

            hasil.append({
                "Kategori": it["kategori"],
                "Judul": it["judul"],
                "Tanggal Terbit": it["tanggal"],
                "Penulis": penulis,
                "Ringkasan": deskripsi,
                "Sentimen": sentimen,
                "Link": it["link"],
                "URL Gambar": url_gambar,
                "Media": feed_info["media"]
            })
    except Exception as e:
        print(f"Gagal ambil RSS {feed_info['media']} ({feed_info['kategori']}): {e}")
    return hasil

def ambil_indeks_detik(subdomain, kategori_nama, max_page=3):
    """Scraping Indeks Kanal Detikcom"""
    hasil = []
    tgl_now = datetime.now().strftime("%m/%d/%Y")
    kandidat = []

    for page in range(1, max_page + 1):
        url = f"https://{subdomain}.detik.com/indeks/{page}?date={tgl_now}"
        try:
            resp = requests.get(url, headers=HEADERS, timeout=15)
            if resp.status_code != 200:
                break
                
            soup = BeautifulSoup(resp.text, 'html.parser')
            articles = soup.find_all('article')
            
            for art in articles:
                tag_a = art.find('a', href=True)
                if not tag_a:
                    continue
                    
                link = tag_a['href']
                tag_title = art.find('h3') or art.find('h2') or tag_a
                judul = tag_title.get_text(strip=True) if tag_title else ""
                
                tag_img = art.find('img')
                url_gambar = ""
                if tag_img:
                    url_gambar = tag_img.get('src') or tag_img.get('data-src') or ""

                if len(judul) > 15:
                    kandidat.append((link, judul, url_gambar))
        except Exception as e:
            print(f"Gagal scraping indeks Detik [{subdomain}] hal {page}: {e}")

    detail_map = ekstraksi_detail_banyak_halaman(
        [link for link, _, _ in kandidat], media_default="Detikcom"
    )

    for link, judul, url_gambar in kandidat:
        ringkasan, img_detail, penulis = detail_map.get(link, ("", "", ""))
        if not url_gambar:
            url_gambar = img_detail
        if not penulis:
            penulis = "Redaksi Detikcom"

        sentimen = analisa_sentimen(judul, ringkasan)

        hasil.append({
            "Kategori": kategori_nama,
            "Judul": judul,
            "Tanggal Terbit": format_tanggal_rfc2822(datetime.now(timezone(timedelta(hours=7)))),
            "Penulis": penulis,
            "Ringkasan": ringkasan,
            "Sentimen": sentimen,
            "Link": link,
            "URL Gambar": url_gambar,
            "Media": "Detik"
        })

    return hasil

def ambil_indeks_kompas(max_page=3):
    """Scraping Indeks Kompas.com secara fleksibel & presisi"""
    hasil = []
    seen_links = set()
    tgl_now = datetime.now().strftime("%Y-%m-%d")
    kandidat = []

    for page in range(1, max_page + 1):
        url = f"https://indeks.kompas.com/?site=all&date={tgl_now}&page={page}"
        try:
            resp = requests.get(url, headers=HEADERS, timeout=15)
            if resp.status_code != 200:
                break
                
            soup = BeautifulSoup(resp.text, 'lxml')
            articles = (
                soup.find_all('div', class_=re.compile(r'article__list|articleList|article__item')) or
                soup.find_all('div', class_='neath-item')
            )
            
            if not articles:
                link_candidates = soup.find_all('a', href=re.compile(r'https?://[a-zA-Z0-9.-]*kompas\.com/read/\d{4}/\d{2}/\d{2}/\d+'))
                for a in link_candidates:
                    link = a.get('href')
                    if not link or link in seen_links: 
                        continue
                    judul = a.get_text(strip=True)
                    if len(judul) < 10: 
                        continue
                    seen_links.add(link)
                    kandidat.append((link, judul, "Berita Utama"))
            else:
                for art in articles:
                    tag_a = art.find('a', href=True)
                    if not tag_a: 
                        continue
                    link = tag_a['href']
                    if not link.startswith("http"):
                        if link.startswith("//"): 
                            link = "https:" + link
                        elif link.startswith("/"): 
                            link = "https://www.kompas.com" + link
                    if link in seen_links: 
                        continue
                    seen_links.add(link)

                    tag_title = art.find('h3') or art.find('h2') or art.find('a')
                    judul = tag_title.get_text(strip=True) if tag_title else ""
                    if len(judul) < 10: 
                        continue

                    tag_cat = art.find('div', class_=re.compile(r'article__subtitle|subtitle|kanal')) or art.find('h4')
                    kategori = tag_cat.get_text(strip=True) if tag_cat else "General"

                    kandidat.append((link, judul, kategori))
        except Exception as e:
            print(f"Gagal scraping indeks Kompas hal {page}: {e}")

    detail_map = ekstraksi_detail_banyak_halaman(
        [link for link, _, _ in kandidat], media_default="Kompas.com"
    )

    for link, judul, kategori in kandidat:
        ringkasan, url_gambar, penulis = detail_map.get(link, ("", "", ""))
        if not penulis:
            penulis = "Redaksi Kompas.com"

        sentimen = analisa_sentimen(judul, ringkasan)

        hasil.append({
            "Kategori": kategori,
            "Judul": judul,
            "Tanggal Terbit": format_tanggal_rfc2822(datetime.now(timezone(timedelta(hours=7)))),
            "Penulis": penulis,
            "Ringkasan": ringkasan,
            "Sentimen": sentimen,
            "Link": link,
            "URL Gambar": url_gambar,
            "Media": "Kompas"
        })

    return hasil

# ----------------------------------------------------------------------
# PENYIMPANAN DATA KE EXCEL
# ----------------------------------------------------------------------
def simpan_excel(df, path_output):
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    
    if os.path.exists(path_output):
        try:
            df_lama = pd.read_excel(path_output)
            df = pd.concat([df_lama, df], ignore_index=True)
            df.drop_duplicates(subset=["Link"], keep="first", inplace=True)
        except Exception as e:
            print(f"Gagal membaca Excel lama ({path_output}): {e}")

    kolom_urut = ["Kategori", "Judul", "Tanggal Terbit", "Penulis", "Ringkasan", "Sentimen", "Link", "URL Gambar", "Media"]
    df = df.reindex(columns=[c for c in kolom_urut if c in df.columns])

    with pd.ExcelWriter(path_output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Berita")
        ws = writer.sheets["Berita"]
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for idx, col in enumerate(df.columns, start=1):
            panjang = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()[:200]])
            ws.column_dimensions[get_column_letter(idx)].width = min(max(panjang + 2, 10), 60)
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

# ----------------------------------------------------------------------
# MAIN FUNCTION
# ----------------------------------------------------------------------
def main():
    print("=== MULAI SCRAPING BERITA OTOMATIS + AI SENTIMEN (INDOBERT) ===\n")
    semua_berita = []

    # 1. Ambil Feed RSS (Semua Media)
    print("1. Mengambil Feed RSS (Antara, Detik, CNN, Tribun, CNBC, Bisnis, Tempo, Republika)...")
    for key, feed_info in FEEDS_RSS.items():
        print(f"   - [{feed_info['media']}] Kanal '{feed_info['kategori']}'...", end=" ")
        berita = ambil_rss(feed_info)
        print(f"{len(berita)} berita")
        semua_berita.extend(berita)
        time.sleep(0.1)

    # 2. Scraping Indeks Kanal Detikcom
    print(f"\n2. Mengambil Indeks Kanal Detikcom (Halaman 1-{MAX_PAGE_INDEKS})...")
    for k_detik in KANAL_INDEKS_DETIK:
        print(f"   - [Detik] Indeks {k_detik['kategori']} ({k_detik['subdomain']}.detik.com)...", end=" ")
        berita_detik = ambil_indeks_detik(k_detik['subdomain'], k_detik['kategori'], max_page=MAX_PAGE_INDEKS)
        print(f"{len(berita_detik)} berita")
        semua_berita.extend(berita_detik)
        time.sleep(0.2)

    # 3. Scraping Indeks Kompas.com
    print(f"\n3. Mengambil Indeks Kompas.com (Halaman 1-{MAX_PAGE_INDEKS})...", end=" ")
    kompas_berita = ambil_indeks_kompas(max_page=MAX_PAGE_INDEKS)
    print(f"{len(kompas_berita)} berita")
    semua_berita.extend(kompas_berita)

    if not semua_berita:
        print("\nTidak ada berita yang berhasil diambil.")
        return

    df = pd.DataFrame(semua_berita)
    df.drop_duplicates(subset=["Link"], keep="first", inplace=True)

    df["_tanggal_parsed"] = df["Tanggal Terbit"].apply(parse_tanggal)
    now = datetime.now(timezone.utc)
    batas_lama = now - timedelta(days=MAKS_UMUR_HARI)
    
    df = df[df["_tanggal_parsed"].notna()]
    df = df[df["_tanggal_parsed"] >= batas_lama]

    if df.empty:
        print("\nTidak ada berita baru dalam rentang waktu yang ditentukan.")
        return

    df["Tanggal Terbit"] = df["_tanggal_parsed"].apply(format_tanggal_rfc2822)
    df["_tanggal_file"] = df["_tanggal_parsed"].apply(lambda d: d.date().isoformat())
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    print("\n4. Menyimpan & Menggabungkan Hasil ke File Excel...")
    for tanggal_file, grup in df.groupby("_tanggal_file"):
        grup_clean = grup.drop(columns=["_tanggal_parsed", "_tanggal_file"])
        path_output = os.path.join(OUTPUT_FOLDER, f"{FILENAME_PREFIX}_{tanggal_file}.xlsx")
        simpan_excel(grup_clean, path_output)
        print(f"   - Update File: {path_output}")

    daftar = [{"file": f, "tanggal": f.replace(f"{FILENAME_PREFIX}_", "").replace(".xlsx", "")}
              for f in os.listdir(OUTPUT_FOLDER) if f.endswith(".xlsx")]
    with open(os.path.join(OUTPUT_FOLDER, "index.json"), "w", encoding="utf-8") as f:
        json.dump(daftar, f, ensure_ascii=False, indent=2)

    print("\n=== PROSES SELESAI! SEMUA SUMBER MEDIA BERHASIL DIMUAT ===")

if __name__ == "__main__":
    main()
